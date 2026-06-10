from __future__ import annotations

import io

from app.core.logging import get_logger
from app.domain.models import FileDescriptor, ParsedDocument, ParsedSection
from app.services.parsers.assets import write_image_asset
from app.services.parsers.base import BaseParser

logger = get_logger(component="xlsx_parser")


class XLSXParser(BaseParser):
    def parse(self, file_descriptor: FileDescriptor) -> ParsedDocument:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        workbook = load_workbook(
            filename=str(file_descriptor.file_path),
            read_only=False,
            data_only=True,
        )

        sections: list[ParsedSection] = []
        try:
            for sheet in workbook.worksheets:
                table_lines: list[str] = []
                for table_name, table in sheet.tables.items():
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                    table_lines.append(f"[TABLE] name={table_name} range={table.ref}")
                    for row_idx in range(min_row, max_row + 1):
                        row_values: list[str] = []
                        for col_idx in range(min_col, max_col + 1):
                            value = sheet.cell(row=row_idx, column=col_idx).value
                            row_values.append(str(value).strip() if value is not None else "")
                        if any(row_values):
                            table_lines.append(" | ".join(row_values))

                if table_lines:
                    sections.append(
                        ParsedSection(
                            text="\n".join(table_lines),
                            section_title=f"{sheet.title} tables",
                        )
                    )

                lines: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        lines.append(" | ".join(values))

                if lines:
                    # Mark the raw sheet grid as a table so the chunker preserves
                    # it intact (the semantic splitter would otherwise break rows
                    # apart) and the UI re-renders it as a table rather than prose.
                    sections.append(
                        ParsedSection(
                            text="[TABLE]\n" + "\n".join(lines),
                            section_title=sheet.title,
                        )
                    )

                image_lines: list[str] = []
                visual_refs: list[dict[str, str]] = []
                for image_index, image in enumerate(getattr(sheet, "_images", []), start=1):
                    anchor = getattr(image, "anchor", None)
                    anchor_text = ""
                    if anchor is not None and hasattr(anchor, "_from"):
                        start = anchor._from
                        anchor_text = f"cell={start.col + 1},{start.row + 1}"

                    name = f"{sheet.title}_img_{image_index}"
                    path = self._write_sheet_image(file_descriptor, image, name)
                    if path:
                        visual_refs.append({"type": "image", "name": name, "path": path})
                    image_lines.append(f"[IMAGE] sheet={sheet.title} {anchor_text}".strip())

                if image_lines:
                    sections.append(
                        ParsedSection(
                            text="\n".join(image_lines),
                            section_title=f"{sheet.title} images",
                            metadata={"visual_refs": visual_refs},
                        )
                    )
        finally:
            workbook.close()

        return ParsedDocument(document_id=file_descriptor.document_id, sections=sections)

    @staticmethod
    def _write_sheet_image(file_descriptor: FileDescriptor, image, name: str) -> str | None:
        """Best-effort extraction of an embedded worksheet image to disk."""
        try:
            data = getattr(image, "_data", None)
            blob: bytes | None = None
            if callable(data):
                blob = data()
            else:
                ref = getattr(image, "ref", None)
                if isinstance(ref, (bytes, bytearray)):
                    blob = bytes(ref)
                elif isinstance(ref, io.BytesIO):
                    blob = ref.getvalue()
            if not blob:
                return None
            ext = (getattr(image, "format", None) or "png").lower()
            return write_image_asset(file_descriptor.file_path, blob, ext, name)
        except Exception as error:
            logger.warning(
                "XLSX image extraction failed",
                file_path=str(file_descriptor.file_path),
                error=str(error),
            )
            return None
